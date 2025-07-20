import os
import logging
import psycopg2
import datetime
import jdatetime
import openpyxl
import pytz
from collections import defaultdict

from dotenv import load_dotenv
from telegram import (
    Update,
    ReplyKeyboardMarkup,
    ReplyKeyboardRemove,
    InputFile,
)
from telegram.ext import (
    Application,
    ContextTypes,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    filters,
)

load_dotenv()
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# --- CONFIG ---
BOT_TOKEN = os.getenv("BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
SUPER_ADMIN = int(os.getenv("SUPER_ADMIN", 0))

(
    GET_USER_MONTH,
    GET_ADMIN_MONTH,
    SET_USER_ID_FOR_NAME,
    SET_NEW_NAME,
) = range(4)

def get_db():
    return psycopg2.connect(DATABASE_URL, sslmode="require")

def ensure_user(user):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT employee_id FROM users WHERE user_id=%s;", (user.id,))
    if cur.fetchone() is None:
        cur.execute("SELECT MAX(employee_id) FROM users;")
        max_id_str = cur.fetchone()[0]
        next_id = int(max_id_str) + 1 if max_id_str and max_id_str.isdigit() else 1
        emp_id = f"{next_id:04d}"

        full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
        display_name = user.first_name or full_name

        cur.execute("""
            INSERT INTO users (user_id, full_name, username, display_name, employee_id)
            VALUES (%s, %s, %s, %s, %s)
            ON CONFLICT (user_id) DO NOTHING;
        """, (user.id, full_name, user.username or '', display_name, emp_id))
        conn.commit()
    cur.close()
    conn.close()

def add_admin(user_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO admins (user_id) VALUES (%s) ON CONFLICT (user_id) DO NOTHING;", (user_id,))
    conn.commit(); cur.close(); conn.close()

def is_admin(user_id):
    if user_id == SUPER_ADMIN:
        return True
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT 1 FROM admins WHERE user_id=%s;", (user_id,))
    ok = cur.fetchone() is not None
    cur.close(); conn.close()
    return ok

def list_users():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT user_id, full_name, username, display_name, employee_id FROM users ORDER BY employee_id;")
    rows = cur.fetchall()
    cur.close(); conn.close()
    return rows

def set_display_name(user_id, new_name):
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE users SET display_name=%s WHERE user_id=%s;", (new_name, user_id))
    conn.commit(); cur.close(); conn.close()

def save_attendance(user_id, status):
    conn = get_db(); cur = conn.cursor()
    now = datetime.datetime.now(pytz.timezone('Asia/Tehran'))
    cur.execute("INSERT INTO attendance (user_id, status, \"timestamp\") VALUES (%s, %s, %s);", (user_id, status, now))
    conn.commit(); cur.close(); conn.close()

def fetch_attendance(user_id=None, start=None, end=None):
    conn = get_db(); cur = conn.cursor()
    q = "SELECT user_id, status, \"timestamp\" FROM attendance"
    params, clauses = [], []
    if user_id is not None:
        clauses.append("user_id=%s"); params.append(user_id)
    if start is not None:
        clauses.append("\"timestamp\">=%s"); params.append(start)
    if end is not None:
        clauses.append("\"timestamp\"<=%s"); params.append(end)
    if clauses:
        q += " WHERE " + " AND ".join(clauses)
    q += " ORDER BY \"timestamp\""
    cur.execute(q, tuple(params))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return rows

def to_shamsi(dateobj):
    s = jdatetime.datetime.fromgregorian(datetime=dateobj.astimezone(pytz.timezone('Asia/Tehran')))
    return s.strftime('%Y/%m/%d'), s.strftime('%H:%M:%S')

def parse_year_month(date_str: str):
    try:
        y, m = map(int, date_str.split('-'))
        if not (1 <= m <= 12): raise ValueError("ماه نامعتبر است")
        iran_tz = pytz.timezone('Asia/Tehran')
        is_shamsi = y < 1500
        if is_shamsi:
            start_s = jdatetime.date(y, m, 1)
            start_g = start_s.togregorian()
            start_dt = iran_tz.localize(datetime.datetime(start_g.year, start_g.month, start_g.day))
            days_in_month = 31 if m <= 6 else (30 if m <= 11 else (29 if not start_s.isleap() else 30))
            end_s = jdatetime.date(y, m, days_in_month)
            end_g = end_s.togregorian()
            end_dt = iran_tz.localize(datetime.datetime(end_g.year, end_g.month, end_g.day, 23, 59, 59))
        else:
            start_dt = iran_tz.localize(datetime.datetime(y, m, 1))
            next_month = m+1 if m < 12 else 1
            next_year = y if m < 12 else y+1
            next_month_start = iran_tz.localize(datetime.datetime(next_year, next_month, 1))
            end_dt = next_month_start - datetime.timedelta(seconds=1)
        return start_dt, end_dt
    except Exception:
        raise ValueError("فرمت تاریخ نامعتبر. مثال: 1403-05 یا 2024-08")

def get_display_name(user_id, users_list):
    for u in users_list:
        if u[0] == user_id:
            return u[3] or u[1] or u[4] or str(user_id)
    return str(user_id)

def main_menu_keyboard(is_admin_user=False):
    kb = [['ثبت ورود', 'ثبت خروج'], ['گزارش روزانه', 'گزارش ماهانه من']]
    if is_admin_user:
        kb.append(['پنل ادمین'])
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def admin_menu_keyboard():
    kb = [['لیست کاربران', 'تعیین نام نمایشی'], ['گزارش روزانه همه', 'گزارش ماهانه همه'], ['دریافت بکاپ اکسل', 'بازگشت به منوی اصلی']]
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def back_keyboard():
    return ReplyKeyboardMarkup([['لغو']], resize_keyboard=True)

async def start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    ensure_user(user)
    if user.id == SUPER_ADMIN:
        add_admin(SUPER_ADMIN)
    admin_status = is_admin(user.id)
    await update.message.reply_text(
        f'سلام {user.first_name}، خوش آمدید!',
        reply_markup=main_menu_keyboard(is_admin_user=admin_status)
    )

async def handle_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    txt = update.message.text
    user = update.effective_user
    ensure_user(user)
    admin_status = is_admin(user.id)

    if txt == 'لغو':
        markup = admin_menu_keyboard() if admin_status else main_menu_keyboard(admin_status)
        await update.message.reply_text("مکالمه لغو شد.", reply_markup=markup)
        return ConversationHandler.END

    if txt == 'ثبت ورود':
        save_attendance(user.id, 'enter')
        await update.message.reply_text("✅ ورود شما ثبت شد.", reply_markup=main_menu_keyboard(admin_status))
    elif txt == 'ثبت خروج':
        save_attendance(user.id, 'exit')
        await update.message.reply_text("✅ خروج شما ثبت شد.", reply_markup=main_menu_keyboard(admin_status))
    elif txt == 'گزارش روزانه':
        await send_daily_report(update, ctx, user.id, all_users=False)
    elif txt == 'گزارش ماهانه من':
        await update.message.reply_text("لطفاً ماه و سال مورد نظر را به فرمت `YYYY-MM` وارد کنید (مثال: `1403-05`).", reply_markup=back_keyboard(), parse_mode='Markdown')
        return GET_USER_MONTH
    elif txt == 'پنل ادمین' and admin_status:
        await update.message.reply_text("شما وارد پنل ادمین شدید.", reply_markup=admin_menu_keyboard())
    elif txt == 'بازگشت به منوی اصلی':
        await update.message.reply_text("بازگشت به منوی اصلی.", reply_markup=main_menu_keyboard(admin_status))
    elif txt == 'لیست کاربران' and admin_status:
        users = list_users()
        lines = [f"👥 **لیست کاربران ({len(users)} نفر)**", "ID - emp_id - نام نمایشی - @username", "---"]
        for u_id, full, uname, dname, emp_id in users:
            lines.append(f"{u_id} - {emp_id} - {dname} - @{uname}")
        await update.message.reply_text('\n'.join(lines), reply_markup=admin_menu_keyboard(), parse_mode='Markdown')
    elif txt == 'تعیین نام نمایشی' and admin_status:
        await update.message.reply_text("لطفاً `ID` عددی کاربر مورد نظر را وارد کنید.", reply_markup=back_keyboard(), parse_mode='Markdown')
        return SET_USER_ID_FOR_NAME
    elif txt == 'گزارش روزانه همه' and admin_status:
        await send_daily_report(update, ctx, user.id, all_users=True)
    elif txt == 'گزارش ماهانه همه' and admin_status:
        await update.message.reply_text("لطفاً ماه و سال مورد نظر را به فرمت `YYYY-MM` وارد کنید (مثال: `1403-05`).", reply_markup=back_keyboard(), parse_mode='Markdown')
        return GET_ADMIN_MONTH
    elif txt == 'دریافت بکاپ اکسل' and admin_status:
        await backup_excel(update, ctx)
    else:
        await update.message.reply_text("دستور شناخته نشد. لطفاً از دکمه‌ها استفاده کنید.", reply_markup=main_menu_keyboard(admin_status))
    return ConversationHandler.END

async def send_daily_report(update, ctx, user_id, all_users=False):
    now = datetime.datetime.now(pytz.timezone('Asia/Tehran'))
    start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    end = now.replace(hour=23, minute=59, second=59)
    title = f"گزارش روزانه تاریخ {jdatetime.date.today().strftime('%Y/%m/%d')}"

    items = fetch_attendance(start=start, end=end) if all_users else fetch_attendance(user_id=user_id, start=start, end=end)
    admin_status = is_admin(user_id)
    reply_markup = admin_menu_keyboard() if all_users else main_menu_keyboard(admin_status)

    if not items:
        await update.message.reply_text("📋 موردی برای امروز ثبت نشده است.", reply_markup=reply_markup)
        return

    users_list = list_users()
    if all_users:
        reports_by_user = defaultdict(list)
        for item in items:
            reports_by_user[item[0]].append(item)
        full_text = [f"**{title}**"]
        for uid, user_items in reports_by_user.items():
            name = get_display_name(uid, users_list)
            full_text.append(f"\n👤 **{name}** ({uid})")
            for it in user_items:
                _, ts = to_shamsi(it[2])
                status_fa = 'ورود' if it[1] == 'enter' else 'خروج'
                full_text.append(f"{ts} - {status_fa}")
        await update.message.reply_text("\n".join(full_text), reply_markup=reply_markup, parse_mode='Markdown')
    else:
        name = get_display_name(user_id, users_list)
        text = [f"**{title} برای {name}**"]
        for it in items:
            _, ts = to_shamsi(it[2])
            status_fa = 'ورود' if it[1] == 'enter' else 'خروج'
            text.append(f"{ts} - {status_fa}")
        await update.message.reply_text("\n".join(text), reply_markup=reply_markup, parse_mode='Markdown')

async def backup_excel(update, ctx):
    await update.message.reply_text("در حال آماده‌سازی فایل بکاپ... لطفاً صبر کنید.")
    path = "backup_all_attendance.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['User ID', 'Employee ID', 'نام نمایشی', 'نام کامل', 'تاریخ شمسی', 'ساعت', 'وضعیت'])

    all_users = {u[0]: u for u in list_users()}
    all_items = fetch_attendance()

    for item in all_items:
        user_data = all_users.get(item[0])
        if user_data:
            shamsi_date, shamsi_time = to_shamsi(item[2])
            ws.append([
                item[0], user_data[4], user_data[3], user_data[1],
                shamsi_date, shamsi_time, 'ورود' if item[1] == 'enter' else 'خروج'
            ])
    wb.save(path)
    await update.message.reply_document(document=InputFile(path), filename=path)
    await update.message.reply_text("✅ بکاپ کامل اکسل ارسال شد.", reply_markup=admin_menu_keyboard())
    os.remove(path)

async def get_user_month_cb(update: Update, ctx):
    if update.message.text == 'لغو':
        admin_status = is_admin(update.effective_user.id)
        markup = main_menu_keyboard(admin_status)
        await update.message.reply_text("مکالمه لغو شد.", reply_markup=markup)
        return ConversationHandler.END
    try:
        start_dt, end_dt = parse_year_month(update.message.text.strip())
        await update.message.reply_text("در حال تهیه گزارش اکسل ماهانه...")
        items = fetch_attendance(user_id=update.effective_user.id, start=start_dt, end=end_dt)
        admin_status = is_admin(update.effective_user.id)
        if not items:
            await update.message.reply_text("موردی در این ماه یافت نشد.", reply_markup=main_menu_keyboard(admin_status))
            return ConversationHandler.END

        path = f"report_{update.effective_user.id}_{start_dt.strftime('%Y%m')}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['تاریخ شمسی', 'ساعت', 'وضعیت'])
        for it in items:
            ds, ts = to_shamsi(it[2])
            ws.append([ds, ts, 'ورود' if it[1] == 'enter' else 'خروج'])
        wb.save(path)
        await update.message.reply_document(document=InputFile(path), filename=f"report_{start_dt.strftime('%Y-%m')}.xlsx")
        await update.message.reply_text("✅ گزارش ماهانه شما ارسال شد.", reply_markup=main_menu_keyboard(admin_status))
        os.remove(path)
        return ConversationHandler.END
    except ValueError as e:
        await update.message.reply_text(str(e), reply_markup=back_keyboard())
        return GET_USER_MONTH

async def get_admin_month_cb(update: Update, ctx):
    if update.message.text == 'لغو':
        await update.message.reply_text("عملیات لغو شد.", reply_markup=admin_menu_keyboard())
        return ConversationHandler.END
    try:
        start_dt, end_dt = parse_year_month(update.message.text.strip())
        await update.message.reply_text("در حال تهیه گزارش اکسل ماهانه برای همه کاربران...")
        items = fetch_attendance(start=start_dt, end=end_dt)
        if not items:
            await update.message.reply_text("موردی در این ماه برای هیچ کاربری یافت نشد.", reply_markup=admin_menu_keyboard())
            return ConversationHandler.END

        path = f"all_report_{start_dt.strftime('%Y%m')}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['User ID', 'Employee ID', 'نام نمایشی', 'تاریخ شمسی', 'ساعت', 'وضعیت'])
        users_map = {u[0]: u for u in list_users()}
        for it in items:
            user_info = users_map.get(it[0])
            if not user_info: continue
            ds, ts = to_shamsi(it[2])
            ws.append([it[0], user_info[4], user_info[3], ds, ts, 'ورود' if it[1] == 'enter' else 'خروج'])
        wb.save(path)
        await update.message.reply_document(document=InputFile(path), filename=f"all_report_{start_dt.strftime('%Y-%m')}.xlsx")
        await update.message.reply_text("✅ گزارش ماهانه همه کاربران ارسال شد.", reply_markup=admin_menu_keyboard())
        os.remove(path)
        return ConversationHandler.END
    except ValueError as e:
        await update.message.reply_text(str(e), reply_markup=back_keyboard())
        return GET_ADMIN_MONTH

async def set_user_id_for_name_cb(update: Update, ctx):
    if update.message.text == 'لغو':
        await update.message.reply_text("عملیات لغو شد.", reply_markup=admin_menu_keyboard())
        return ConversationHandler.END
    try:
        user_id_to_edit = int(update.message.text.strip())
        ctx.user_data['user_id_to_edit'] = user_id_to_edit
        await update.message.reply_text("لطفاً نام نمایشی جدید را وارد کنید.", reply_markup=back_keyboard())
        return SET_NEW_NAME
    except (ValueError, TypeError):
        await update.message.reply_text("ID وارد شده معتبر نیست. لطفاً یک ID عددی ارسال کنید.", reply_markup=back_keyboard())
        return SET_USER_ID_FOR_NAME

async def set_new_name_cb(update: Update, ctx):
    if update.message.text == 'لغو':
        await update.message.reply_text("عملیات لغو شد.", reply_markup=admin_menu_keyboard())
        return ConversationHandler.END
    new_name = update.message.text.strip()
    user_id_to_edit = ctx.user_data.pop('user_id_to_edit', None)
    if not user_id_to_edit:
        await update.message.reply_text("خطا! لطفاً مجدداً از پنل ادمین اقدام کنید.", reply_markup=admin_menu_keyboard())
        return ConversationHandler.END
    set_display_name(user_id_to_edit, new_name)
    await update.message.reply_text(f"✅ نام نمایشی کاربر {user_id_to_edit} به {new_name} تغییر یافت.", reply_markup=admin_menu_keyboard())
    return ConversationHandler.END

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    conv_handler = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text)],
        states={
            GET_USER_MONTH: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_user_month_cb)],
            GET_ADMIN_MONTH: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_admin_month_cb)],
            SET_USER_ID_FOR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, set_user_id_for_name_cb)],
            SET_NEW_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, set_new_name_cb)],
        },
        fallbacks=[]
    )
    app.add_handler(CommandHandler("start", start))
    app.add_handler(conv_handler)
    logging.info("Bot Started...")
    app.run_polling()

if __name__ == "__main__":
    main()
